# -*- coding: utf-8 -*-
"""
Created on Fri Apr 20 16:42:28 2018

@author: ZyuKyuman
"""
import json
import requests
from openpyxl import Workbook
from openpyxl import load_workbook

myPlanKey = ['fundName','fundCode','myNav','myUnit']#基金名称，神秘代码，净值，份额

wb = load_workbook('myPlan.xlsx',data_only=True)
myPlanKeyLocation = {'fundName': '', 'fundCode': '', 'myNav': '', 'myUnit': ''} 
sheet = wb.worksheets[0]

#get the max number of column and row
max_column = 1
max_row = 1 
index = 1
while(True):
    key = sheet.cell(row=1, column=index)
    if(key.value!=None):
        index += 1
    else :
        max_column = index - 1
        break;
index = 1
while(True):
    key = sheet.cell(row=index, column=1)
    if(key.value!=None):
        index += 1
    else :
        max_row = index - 1 
        break;

#find the location of my plan's key
index = max_column
while(True):
    key = sheet.cell(row=1, column=index)
    index = index - 1
    if(key.value in myPlanKeyLocation):  
        myPlanKeyLocation[key.value] = key.column
    if(index<=0):
        print(myPlanKeyLocation)
        break;

#{'fundName': 'A', 'fundCode': 'B', 'myNav': 'G', 'myUnit': 'H'}
#get my fund nav    
fundName = []
for cell in sheet[myPlanKeyLocation['fundName']]:
    fundName.append(cell.value)
#print(fundName)

fundCode = []
for cell in sheet[myPlanKeyLocation['fundCode']]:
    fundCode.append(cell.value)
#print(fundCode)

myNav = []
for cell in sheet[myPlanKeyLocation['myNav']]:
    myNav.append(cell.value)
#print(myNav)

myUnit = []
for cell in sheet[myPlanKeyLocation['myUnit']]:
    myUnit.append(cell.value)
#print(myUnit)


#start to get etf data
url = 'https://qieman.com/pmdj/v2/long-win/plan';
url1 = 'https://qieman.com/longwin/detail';
header = {
        'Accept': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Host': 'qieman.com',
        'If-None-Match': 'W/"31cb-1d8SUjBTdgt4PlT8sVjcB5GCI+g"',
        'Referer': 'https://qieman.com/longwin/detail',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36',
        'x-aid': 'GA1.2.564496475.1524208568',
        'x-request-id': 'albus.7BD04909C1CE3D30440A',
        'x-sign': '1541311881027B5FE615F372625E5693E138F5C8CAD72'
}
html = requests.get(url,headers=header)
jason = html.json()['composition']

pre_title = ['A','B','C','D','E','F','G','H','I','J','K','L','M']
myUnitValues = {'中证红利':1.1476,'中证环保':0.6982,'全指医药':0.788,'建信500':2.1745,
          '海外收益债':1.1829,'证券公司':0.988,'中证传媒':0.8968,'中证500':0.6195,
          '富国300':1.8041,'兴全转债':1.0375}
myplanUnits = {'中证红利':4,'中证环保':8,'全指医药':1,'建信500':9,
          '海外收益债':3,'证券公司':2,'中证传媒':3,'中证500':3,'富国300':1,
          '兴全转债':1}

wb = load_workbook('etf.xlsx')
#wb = Workbook()
sheet = wb.worksheets[0]

titlt_name = {'A1':"fundCode",'B1':"fundName",'C1':"etfUnit",'D1':"percent",'E1':"unitNav",
       'F1':"etfNav",'G1':'etfProfit','H1':'myUnit','I1':'myNav','J1':'distanceWithETF',
       'K1':'myProfit','L1':'Total'}

for x in range(1,len(pre_title)):
    s = pre_title[x]+str(1)
    sheet[s] = titlt_name.get(s);
    
for x in range(0,len(jason)):
    code        = (jason[x].get('fund'))['fundCode'];
    nav         = jason[x].get('nav');
    percent     = jason[x].get('percent');
    planUnit    = jason[x].get('planUnit');
    profit      = jason[x].get('profit');
    unitValue   = jason[x].get('unitValue');
    variety     = jason[x].get('variety');
    myUnitValue = myUnitValues.get(variety,0)
    myplanUnit  = myplanUnits.get(variety,0)
    
    if myplanUnit == 0:
        profitWithETF = 0;
        myProfit = 0;
        total = 0;
    else :
        profitWithETF = (unitValue-myUnitValue)/myUnitValue;
        myProfit = (nav-myUnitValue)/myUnitValue;
        total = myplanUnit*500*(1+myProfit);

    jason_value = [code,variety,planUnit,percent,nav,unitValue,profit,
                   myplanUnit,myUnitValue,profitWithETF,myProfit,total]
    print(x,jason_value)
    for y in range(0,len(jason_value)):
        sheet[pre_title[y]+str(x+2)] = jason_value[y];
wb.save('etf.xlsx')