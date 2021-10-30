
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 20 16:42:28 2018

@author: JuCie
"""

#import urllib
import requests
#from selenium import webdriver
#from selenium.webdriver.common.action_chains import ActionChains
#import json
from openpyxl import Workbook
from openpyxl import load_workbook

url = 'https://qieman.com/pmdj/v2/long-win/plan';
url1 = 'https://qieman.com/longwin/detail';
header = {
        'Accept': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Host': 'qieman.com',
        'If-None-Match': 'W/"31cb-1d8SUjBTdgt4PlT8sVjcB5GCI+g"',
        'Referer': 'https://qieman.com/longwin/detail',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.117 Safari/537.36',
        'x-aid': 'GA1.2.564496475.1524208568',
        'x-request-id': 'albus.9AEE35901E6B282FDE10',
        'x-sign': '1539520744686B8092F2B9D14CF3D0E8CC578DC0B4C7F'
}
html = requests.get(url,headers=header);
jason = html.json()['composition'];
pre_title = ['A','B','C','D','E','F','G','H','I','J','K','L','M']
myUnitValues = {'中证红利':1.1476,'中证环保':0.6982,'全指医药':0.788,'建信500':2.1745,
          '海外收益债':1.1829,'证券公司':0.988,'中证传媒':0.8968,'中证500':0.6195,
          '富国300':1.8041,'兴全转债':1.0375}

myplanUnits = {'中证红利':4,'中证环保':8,'全指医药':1,'建信500':9,
          '海外收益债':3,'证券公司':2,'中证传媒':3,'中证500':3,'富国300':1,
          '兴全转债':1}
wb = load_workbook('etf.xlsx')
#wb = Workbook()
sheet = wb.worksheets[0]

titlt_name = {'A1':"编号",'B1':"品种",'C1':"份数",'D1':"占比",'E1':"单位净值",
       'F1':"每份净值",'G1':'收益','H1':'我的份数','I1':'我的净值','J1':'e大收益差距',
       'K1':'收益率','L1':'总额'}

for x in range(1,len(pre_title)):
    s = pre_title[x]+str(1)
    sheet[s] = titlt_name.get(s);
    
for x in range(0,len(jason)):
    nav = jason[x].get('nav');
    percent = jason[x].get('percent');
    planUnit = jason[x].get('planUnit');
    profit = jason[x].get('profit');
    unitValue = jason[x].get('unitValue');
    variety = jason[x].get('variety');
    myUnitValue = myUnitValues.get(variety,0)
    myplanUnit = myplanUnits.get(variety,0)
    
    if myplanUnit == 0:
        profitWithETF = 0;
        myProfit = 0;
        total = 0;
    else :
        profitWithETF = (unitValue-myUnitValue)/myUnitValue;
        myProfit = (nav-myUnitValue)/myUnitValue;
        total = myplanUnit*500*(1+myProfit);

    jason_value = [x+1,variety,planUnit,percent,nav,unitValue,profit,
                   myplanUnit,myUnitValue,profitWithETF,myProfit,total]
    print(x,jason_value)
    for y in range(0,len(jason_value)):
        sheet[pre_title[y]+str(x+2)] = jason_value[y];
wb.save('etf.xlsx')
